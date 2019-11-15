import requests
from selenium import webdriver
from lxml import etree
import re
import openpyxl
from openpyxl import load_workbook
import datetime
import pymysql
from apscheduler.schedulers.blocking import BlockingScheduler

def get_pages():
	path = "/usr/local/bin/chromedriver" # chromedriver完整路径，path是重点
	global browser 
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--no-sandbox')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument('--disable-dev-shm-usage')

	browser = webdriver.Chrome(path,chrome_options=chrome_options)
	browser.get('https://auction.jd.com/sifa_list.html?childrenCateId=12730')
	input_first = browser.find_element_by_xpath('/html/body/div[6]/div/div[2]/div[3]/div/div[2]/div/ul/li[3]')
	input_first.click()
	# next_pages = browser.find_element_by_xpath('')
	# next_pages.click()
	time_list = get_times()
	html = browser.page_source
	urls = get_one_page(html)
	list_out = add_set(urls)
	print(list_out,"今日结束")
	create_excel()
	for url in list_out:
		new_url = "https:"+url
		list1 = get_text(new_url)
	# 	insert_excel(list1)
	# 	save_mysql(list1,time_list,new_url)
	# browser.close()

def get_one_page(html):
	html = etree.HTML(html)
	urls = html.xpath('/html/body/div[6]/div/div[4]/ul/li/a/@href')
	return urls

def get_text(new_url):
	gets_url = [new_url]
	browser.get(new_url)
	html1 = browser.page_source
	html = etree.HTML(html1)
	#标题
	title = html.xpath('//*[@id="root"]/div/div[2]/div[1]/div[2]/div[1]/text()')
	#起拍价
	first_price = html.xpath('//*[@id="root"]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[2]/em/text()')
	#保证金
	caution_money = html.xpath('//*[@id="root"]/div/div[2]/div[1]/div[2]/div[3]/div[3]/div/div[1]/ul/li[3]/em/text()')
	if not caution_money or caution_money[0] == '5分钟/次':      
		caution_money = ['无']

	#评估价
	evaluation_price = html.xpath('//*[@id="root"]/div/div[2]/div[1]/div[2]/div[3]/div[3]/div/div[1]/ul/li[1]/em/text()')
	if not evaluation_price:
		evaluation_price = ['无']
	#加价幅度
	Price_increase = html.xpath('//*[@id="root"]/div/div[2]/div[1]/div[2]/div[3]/div[3]/div/div[1]/ul/li[2]/em/text()')
	#处理单位
	disposal_place = html.xpath('//*[@id="root"]/div/div[2]/div[1]/div[2]/div[2]/em/text()')
	place = re.findall('(.*?市|.*?县|.*?自治州|\w.+).*?',disposal_place[0],re.S)
	#面积
	# area = re.findall('<span>(\d+\.\d+)</span>',html1,re.S)
	# if not area:
	# 	area = ['无']
	list1 =title+gets_url+first_price+caution_money+evaluation_price+Price_increase+place
	print(list1)
	return list1

#获取时间
def get_times():
	today = datetime.date.today()
	yesterday = datetime.date.today() + datetime.timedelta(-1)
	time_list = [today,yesterday]
	return time_list

#保存进mysql
def save_mysql(list,time_list,new_url):
	db = pymysql.connect(host = 'localhost',user = 'root',password = '你的密码' , port=3306 ,db ='Scrapy')
	cursor = db.cursor()
	print("连接成功")
	try:
		cursor.execute("alter table sf_taobao auto_increment=1")
		cursor.execute("insert into sf_taobao\
					(address,url,first_price,caution_money,evaluation_price,Price_increase,disposal_place,date)\
					 Values('%s','%s','%s','%s','%s','%s','%s','%s')" %(list[0],new_url,list[2],list[3],list[4],list[5],list[6],str(time_list[0])))
		db.commit()
		print("插入成功")
	except Exception as e:
		print("插入失败",e)
		db.rollback()
	db.close

#创建列表
def create_excel():
	mywb = openpyxl.Workbook()
	sheet = mywb.active
	sheet.title = '淘宝司法拍卖土地每日数据'
	title = ['标题','链接','起拍价','保证金','评估价','加价幅度','处理地']
	zm = ['A','B','C','D','E','F','G']
	for i in range(len(title)):
		sheet[str(zm[i])+'1'] =title[i]
	mywb.save('淘宝司法拍卖土地每日数据.xlsx')
#插入excel保存
def insert_excel(list):
	try:
		wb = load_workbook('淘宝司法拍卖土地每日数据.xlsx')
		sheet = wb.get_sheet_by_name('淘宝司法拍卖土地每日数据') # 获得指定名称页
		if len(list) == 8:
			sheet.append(list[:-1])
		elif len(list) == 9:
			sheet.append(list[:-2])
		elif len(list) == 10:
			sheet.append(list[:-3])
		else:
			sheet.append(list)
		wb.save('淘宝司法拍卖土地每日数据.xlsx')
	except Exception as e:
		print("添加失败",e)
#去重
def add_set(urls):
	list_setA = set(urls)
	list3 =[]
	with open('sf_jd.txt') as flle_read:
		lines = flle_read.readlines()
		for line in lines:
			line=line.strip("\n")
			list3.append(str(line))
	list_setB = set(list3)
	list_out = list_setA.difference(list_setB)
	list_out = list(list_out)
	with open('sf_jd.txt','w+') as file_object:
		for test in urls:
			file_object.write(test+'\n')
	return list_out

def main():
	get_pages()

#定时器
def scheduler():
	try:
		scheduler = BlockingScheduler()
		scheduler.add_job(main, 'cron', hour=8, minute=0 ,misfire_grace_time=3600)
		print("正在运行......")
		scheduler.start()
	except Exception as e:
		print("出错，跳过"+e)




# scheduler()
#测试时直接执行main
main()
