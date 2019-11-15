# -*- coding: utf-8 -*-  

import requests
from lxml import etree
import re
import json
import pymysql
import openpyxl
from openpyxl import load_workbook
import datetime
from apscheduler.schedulers.blocking import BlockingScheduler
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication 
import os

def get_pages(first_url,headers):
	r = requests.get(first_url,headers = headers)
	html = r.text
	url = re.findall('itemUrl.*?//(.*?)".*?status',html,re.S)
	return url
	
def get_html(url,headers):
	try:
		r = requests.get(url,headers = headers)
		html = etree.HTML(r.text)
		return html
	except Exception as e:
		print("页面获取失败"+e)
def get_text(html,gets_url):
	#地址
	title = html.xpath('//*[@id="page"]/div[4]/div/div/h1/text()')
	title = [x.strip() for x in title if x.strip() != '']
	#起拍价
	first_price = html.xpath('//*[@id="J_HoverShow"]/tr[1]/td[1]/span[2]/span/text()')
	#保证金
	caution_money = html.xpath('//*[@id="J_HoverShow"]/tr[2]/td[1]/span[2]/span/text()')
	#评估价
	evaluation_price = html.xpath('//*[@id="J_HoverShow"]/tr[3]/td[1]/span[2]/span/text()')
	if not evaluation_price:
		evaluation_price = ['无']
	#加价幅度
	Price_increase = html.xpath('//*[@id="J_HoverShow"]/tr[1]/td[2]/span[2]/span/text()')
	#处理单位
	disposal_place = html.xpath('//*[@id="page"]/div[4]/div/div/div[3]/div[2]/div[1]/span[2]/a/text()')
	place = re.findall('(.*?市|.*?县|.*?自治州|\w.+).*?',disposal_place[0],re.S)
	list1 =title+gets_url+first_price+caution_money+evaluation_price+Price_increase+place
	# area1 = html.xpath('//*/text()')
	# area = re.findall('<span.*?>(\d+\.\d+)</span>',area1[0],re.S)
	# if not area:
	# 	area = ['无']
	# print(area)
	return list1
#保存进mysql
def save_mysql(list,time_list,gets_url):
	db = pymysql.connect(host = 'localhost',user = 'root',password = '你的密码' , port=3306 ,db ='crawl')
	cursor = db.cursor()
	print("连接成功")
	try:
		cursor.execute("alter table sf_taobao auto_increment=1")
		cursor.execute("insert into sf_taobao\
					(address,url,first_price,caution_money,evaluation_price,Price_increase,disposal_place,date)\
					 Values('%s','%s','%s','%s','%s','%s','%s','%s')" %(list[0],gets_url[0],list[2],list[3],list[4],list[5],list[6],str(time_list[0])))
		db.commit()
		print("插入成功")
	except Exception as e:
		print("插入失败",e)
		db.rollback()
	db.close
#创建列表
def create_excel():
	mywb = openpyxl.Workbook()
	sheet = mywb.active
	sheet.title = '淘宝司法拍卖土地每日数据'
	title = ['标题','链接','起拍价','保证金','评估价','加价幅度','处理地']
	zm = ['A','B','C','D','E','F','G']
	for i in range(len(title)):
		sheet[str(zm[i])+'1'] =title[i]
	mywb.save('淘宝司法拍卖土地每日数据.xlsx')
#插入excel保存
def insert_excel(list):
	print(list)
	try:
		wb = load_workbook('淘宝司法拍卖土地每日数据.xlsx')
		sheet = wb.get_sheet_by_name('淘宝司法拍卖土地每日数据') # 获得指定名称页
		if len(list) == 8:
			sheet.append(list[:-1])
		else:
			sheet.append(list)
		wb.save('淘宝司法拍卖土地每日数据.xlsx')
	except Exception as e:
		
		print("添加失败",e)
#获取时间
def get_times():
	today = datetime.date.today()
	yesterday = datetime.date.today() + datetime.timedelta(-1)
	time_list = [today,yesterday]
	return time_list
#定时器
def scheduler():
	try:
		scheduler = BlockingScheduler()
		scheduler.add_job(main, 'cron', hour=8, minute=25 ,misfire_grace_time=3600)
		print("正在运行......")
		scheduler.start()
	except Exception as e:
		print("出错，跳过"+e)
#获取页数
def get_page_nums(url,headers):
	try:
		r = requests.get(url,headers = headers)
		html = etree.HTML(r.text)
		nums = html.xpath('/html/body/div[3]/div[4]/span[4]/em/text()')
		return nums[0]
	except Exception as e:
		print("今天没有新的标的物产生"+e)
#遍历生成一个链接列表
def get_urltext(url,page_nums,headers):
	list_url = []
	for i in range(1,int(page_nums)+1):
		first_url = url +'&page='+str(i)
		pages = get_pages(first_url,headers)
		list_url +=pages
	return list_url
def add_set(list_url):
	list_setA = set(list_url)
	list3 =[]
	with open('sf_taobao.txt') as flle_read:
		lines = flle_read.readlines()
		for line in lines:
			line=line.strip("\n")
			list3.append(str(line))
	list_setB = set(list3)
	list_out = list_setA.difference(list_setB)
	list_out = list(list_out)
	with open('sf_taobao.txt','w+') as file_object:
		for test in list_url:
			file_object.write(test+'\n')
	return list_out
#发送邮件
def send_mail(time_list):
        fromaddr = '你的邮箱'
        password = '你的邮箱码'
        toaddrs = ['']


        content = 'Dear all,\n\t附件为今日淘宝司法拍卖土地信息，请查收\n\n'
        textApart = MIMEText(content)

    
 
        zipFile = '淘宝司法拍卖土地每日数据.xlsx'
        zipApart = MIMEApplication(open(zipFile, 'rb').read())
        zipApart.add_header('Content-Disposition', 'attachment', filename=zipFile)
 
        m = MIMEMultipart()
        m.attach(textApart)
        m.attach(zipApart)
        m['Subject'] = '淘宝司法拍卖土地'+str(time_list[0])+'数据'
        m['From'] = ""

 
        try:
            server = smtplib.SMTP('smtp.qq.com')
            server.login(fromaddr,password)
            server.sendmail(fromaddr, toaddrs, m.as_string())
            print('success')
            server.quit()
        except smtplib.SMTPException as e:
            print('error:',e) #打印错误	
#结束后删除程序
def deleteexcel():
	try:
		os.remove("淘宝司法拍卖土地每日数据.xlsx")
		print("文件删除完毕")
	except(FileNotFoundError):
		print("文件不存在")
def main():
	time_list = get_times()
	headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
    }
	url = 'https://sf.taobao.com/item_list.htm?spm=a213w.7398504.pagination.1.2e955fe8lHANme&category=50025970&auction_source=0&sorder=1&st_param=-1&auction_start_seg=-1'
	page_nums = get_page_nums(url,headers)
	print(page_nums)
	# print("列表",get_urltext(url,page_nums,headers))

	list_url = get_urltext(url,page_nums,headers)
	date_nums = len(list_url)
	print(date_nums)
	list_out = add_set(list_url)
	print(list_out,"今日结束")

	create_excel()
	for page in list_out:
		new_url = "https://"+page
		print(new_url)
		gets_url = new_url.split(' ')
		html = get_html(new_url,headers)
		list =get_text(html,gets_url)
	# 	print(list)
		insert_excel(list)
		# save_mysql(list,time_list,gets_url)
	send_mail(time_list)
	deleteexcel()
# scheduler()
#测试时直接执行main
main()
